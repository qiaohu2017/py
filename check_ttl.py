#!python2.7
# coding:utf-8  #must be line 1
# ttl家长名字优先比较 170507
# cmb 多个记录比如1270，20要调整一下
from openpyxl import Workbook
from openpyxl import load_workbook
import sys

reload(sys) 
sys.setdefaultencoding('gb18030') #for UnicodeDecodeError: 'ascii' codec can't decode byte 0xa1 in position 0: ordinal not in range(128)

orderType = ["新订","续订","复活"]

wb = load_workbook("TTL.xlsx")
ws_ttl = wb.get_sheet_by_name("ttl")
ws_cmb = wb.get_sheet_by_name("cmb")
#ws_test = wb.get_sheet_by_name("test")

#get  value
MaterNameDict = {}
ParentNameDict = {}
rowCounter  = 0
for row in tuple(ws_ttl.rows):
  rowCounter = rowCounter + 1
  #payerName = row[4].value    
  #payerNameU = unicode(payerName)
  #if (payerNameU.encode('utf8') in orderType ):
  #   payerName = row[5].value  
  MaterName = row[5].value  
  if not MaterName  :
      continue     
  print("ttl get MaterName:%s "% MaterName)
  MaterNameDict[MaterName] = rowCounter

  ParentName = row[11].value  
  if not ParentName  :
      continue     
  print("ttl get ParentName:%s "% ParentName)
  ParentNameDict[ParentName] = rowCounter

  ws_ttl["A"+str(rowCounter)] = ""

rowCounter  = 0
for row in tuple(ws_cmb.rows):
   rowCounter = rowCounter + 1
   paymentDesc = row[6].value  
   paymentPrice = row[3].value    
   print("cmb get payment price: %s, description: %s "% (paymentPrice, paymentDesc))
   if not paymentDesc  :
      continue     

   ttl_pos = ""
   for payerName in ParentNameDict.keys():
       if payerName in paymentDesc:
          ttl_pos = "A"+ str(ParentNameDict[payerName])
          break
   if (ttl_pos == ""):
     for payerName in MaterNameDict.keys():
       if payerName in paymentDesc:
          ttl_pos = "A"+ str(MaterNameDict[payerName])
          break

   if (ttl_pos == ""):   #not find
      print("cmb record failed find payer: %s"% (paymentDesc))
      continue           
   
   cmb_pos = "H"   + str(rowCounter)
   ws_cmb[cmb_pos] = payerName      
   print("cmb writed  %s. payerName:%s "% (cmb_pos, payerName))       
   #ws_test.cell(rowCounter, 1).value = payerName
   
   ws_ttl[ttl_pos] = paymentPrice
   print("ttl writed  %s. paymentPrice:%s "% (ttl_pos, paymentPrice)) 
  
 
     
 
  
wb.save("TTL.xlsx")
 
