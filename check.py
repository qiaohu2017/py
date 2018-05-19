# coding:utf-8  #must be line 1
# read order.xlsx to db 
# test for git
from openpyxl import Workbook
from openpyxl import load_workbook
import sys

reload(sys) 
sys.setdefaultencoding('gb18030') #for UnicodeDecodeError: 'ascii' codec can't decode byte 0xa1 in position 0: ordinal not in range(128)

orderType = ["新订","续订","复活"]

wb = load_workbook("TTL.xlsx")
ws_ttl = wb.get_sheet_by_name("ttl")
ws_cmb = wb.get_sheet_by_name("cmb")
ws_test = wb.get_sheet_by_name("test")

#get  value
payerNameList = []
for row in tuple(ws_ttl.rows):
  #payerName = row[4].value    
  #payerNameU = unicode(payerName)
  #if (payerNameU.encode('utf8') in orderType ):
  #   payerName = row[5].value  
  MaterName = row[5].value  
  if not MaterName  :
      continue     
  print("get MaterName:%s "% MaterName)
  payerNameList.append(MaterName)

  ParentName = row[11].value  
  if not ParentName  :
      continue     
  print("get ParentName:%s "% ParentName)
  payerNameList.append(ParentName)

rowCounter  = 0
for row in tuple(ws_cmb.rows):
   rowCounter = rowCounter + 1
   paymentDesc = row[6].value  
   paymentPrice = row[3].value    
   print("get payment price: %s, description: %s "% (paymentPrice, paymentDesc))
   if not paymentDesc  :
      continue     

   bfind = False
   for payerName in payerNameList:
       if payerName in paymentDesc:
               pos = "H"   + str(rowCounter)
               #ws_test.cell(rowCounter, 1).value = payerName
               ws_cmb[pos] = payerName
               print("found writed %s "% (payerName)) 
               bfind = True
               break
   if (not bfind):
     print("failed %s %s"% (rowCounter,paymentDesc))
 
  
wb.save("TTL.xlsx")
 
